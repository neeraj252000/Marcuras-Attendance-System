# -*- coding: utf-8 -*-
"""
Notes:
    Clear ID_text_box after successful submission

@author: Neeraj
"""
from tkinter import *
#from tkinter.ttk import *
from tkinter import Text
import tkinter.filedialog as file
import cv2
import numpy as np
from PIL import Image
from PIL import ImageTk
import numpy 
import tkinter.messagebox

import PIL.Image, PIL.ImageTk
import time
#import face_recognition
#import pickle
import openpyxl
import datetime
import csv
import os

data_length = 0
max_data = 0
img1 = numpy.ndarray
img2 = numpy.ndarray

folder_path = os.path.abspath('.')
#encoding_file = folder_path + "\\encodings.pickle"
#image_read_path = folder_path + "\\stored_photo.jpg"
names_file_path = folder_path + "\\names.csv"
excel_file_path = folder_path + "\\myxl.xlsx"


def Display_Data_Size():
    
    global data_length
    data = text_entry.get("1.0", "end-1c") 
    data_length = len(data)           #+1 for the delimeter
    if str(data).isdigit() or str(data) == "":
        data_label['fg'] = 'black'
    else:
        data_label['fg'] = 'red'
    data_label['text'] = "Data Size: " + str(data_length)
    root.after(1, Display_Data_Size)
    #print(data_length)
    

def get_detected_ID(self,name):
    #Importing User data CSV file
    reader = csv.reader(open(names_file_path))
    #Creating dictionary with names and IDs
    result = {}
    for row in reader:
        result[row[1]] = row[0]
        print(row[1])

    ID = result.get(name)
    #Typecasting to int
    #handling error if name not found in dictionary - returning 0
    try:
        return(int(ID))
    except TypeError:
        return 0
    
def Upload_IMG():
    '''
    global max_data
    global img1
    #global img, image, FilePath
    FilePath = file.askopenfilename()
    
    img = Image.open(FilePath)
    img1 = cv2.cvtColor(numpy.array(img), cv2.COLOR_RGB2BGR)
    resized_img = img.resize(Dim(img, height=300))
    canvas.image = ImageTk.PhotoImage(resized_img)
    canvas.create_image(200,150, image=canvas.image, anchor='center')
    
    max_data = img.size[0]*img.size[1]*3//8
    print(img.size)
    max_data_label['text'] = "Max Data: " + str(max_data)
    '''
    return 0

    
def Dim(image, width = None, height = None):
    dim = None
    w, h = image.size
    
    if width is None and height is None:
        return image
    
    elif height is None:
        r = width/float(w) #ratio of new width to old width.
        dim = (width, int(h*r))
        
    elif width is None:
        r = height/float(h)
        dim = (int(w * r), height)
        
    return dim

def confirm(user_name, ID, mode):
        clear_text()
        confirm_window= Tk()
        confirm_window.title("Steganography")
        confirm_window.geometry("400x300")
        confirm_window.configure(background='#474747')
        confirm_window.resizable(width = False, height=False)
        
        def on_closing():
            submit_ID_button.config(state="normal")
            confirm_window.destroy()
        
        confirm_window.protocol("WM_DELETE_WINDOW", on_closing)
        confirm_window.mainloop()
        
        
def Encode():
    try:    
        reader = csv.reader(open(names_file_path))
    except PermissionError:
        tkinter.messagebox.showerror("Error : 3", "Permission Error! \n Close the opened Excel Files and try again")
        return 0
    
    #Checking for ID in database
    count = 0
    user_name = "null"
    for row in reader:
        #count = count + 1
        #print(row[0])
        #print(count)
        ID = text_entry.get("1.0", "end-1c") 
        if(row[0] == str(ID)):
            print(row[1])
            user_name = row[1]
            break
    if user_name == "null":
        tkinter.messagebox.showerror("Error 1", "Invalid ID! \n ID not found.")
        clear_text()
    else:
        now = datetime.datetime.now()
        current_time = now.strftime("%H:%M:%S")
        #Get date as integer
        date = now.day
        date = int(date)
        
        #Get month as integer 
        month = now.month
        month = int(month)

        #typecasting ID as integer
        ID = int(ID)
        
        # Accessing Attendance Excell
        try:
            x1 = openpyxl.load_workbook(excel_file_path)
            s1 = x1['Sheet'+ str(month)]
            i = 0
        except PermissionError:
            print("Permission Error!")
            tkinter.messagebox.showerror("Error : 3", "Permission Error! \n Close the opened Excel Files and try again")
            return 0
        
        while True:
            i = i+1
            #print(s1.cell(row = i,column = 1).value)
            if(s1.cell(row = i,column = 1).value == str(ID)):
                print(s1.cell(row = i,column = 1).value)
                break
            if(str(s1.cell(row = i,column = 1).value) == 'None' and i >1):
                tkinter.messagebox.showerror("Error : 2", "Found invalid ID field in Attendance Sheet")
                #print("Error: ID field in entry sheet is blank or ID not found")
                break
        
        cell_check = s1.cell(row = i,column = date*3)
        cell_check = str(cell_check.value) 
        
        if(cell_check == "None"):
            s1.cell(row = i,column = date*3).value = current_time
            submit_ID_button["state"] = DISABLED
            confirm(user_name, ID, "in")
           
        else:
            s1.cell(row = i,column = date*3 + 1).value = current_time
            submit_ID_button["state"] = DISABLED
            confirm(user_name, ID, "out")
        
        try:
            x1.save(excel_file_path)
        except PermissionError:
            print("Permission Error!")
            tkinter.messagebox.showerror("Error : 3", "Permission Error! \n Close the opened Excel Files and try again")
            return 0
        # Condition for 2 factor verification - ID in database matches that in attendance sheet
        #if (ID == int(s1.cell(row = i,column = 1).value)):
        
        
        
    
    
    

def get_frame():
    if vid.isOpened():
        ret, frame = vid.read()
        if ret:
            # Return a boolean success flag and the current frame converted to BGR
            return (ret, cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
        else:
            return (ret, None)
    else:
        return (ret, None)

def update():
    # Get a frame from the video sources
    ret, frame = vid.read()

    photo = PIL.ImageTk.PhotoImage(image = PIL.Image.fromarray(frame))
    canvas.create_image(0, 0, image = photo, anchor = 'nw')

    after(10, update)
    
def submit_ID():
    return 0
    

def clear_text():
    text_entry.delete('1.0', END)
    
root = Tk()
root.title("Attendance System")
root.geometry("1200x600")
root.configure(background='#44FE97') 
root.resizable(width = False, height=False)

title = Label(root, text="Marcuras Attendance Tool", bg='#44FC97',
              fg='#000000', font=("Alegreya Sans Bold", 25))
title.pack()

image_label = Label(root, text="Camera: ", bg='#44FC97',
                   fg='#000000', font=("Alegreya Sans Bold", 15))
image_label.pack()
image_label.place(anchor = "w", relx = 0.05, y = 120)

canvas = Canvas(root, bg = 'white', height = 300, width = 400)
canvas.pack()
canvas.place(anchor='center', relx = 0.25, y = 300)

upload_button = Button(root, text="Click!",
                       font=("Alegreya Sans Bold", 10),
                       fg = '#000000', bg = 'white',
                       command = Upload_IMG)
upload_button.pack()
upload_button.place(anchor = "center", relx = 0.25, y = 500,
                    height = 30, width = 90)

enter_text_label = Label(root, text="Enter Employee ID: ", bg='#44FC97',
                         fg='#000000', font=("Alegreya Sans Bold", 15))
enter_text_label.pack()
enter_text_label.place(anchor = "w", relx = 0.6, y = 120)

text_entry = Text(root, width = 20, height = 2, font=("Alegreya Sans", 15))
text_entry.pack()
text_entry.place(anchor = "w", relx = 0.6, y = 180)


submit_ID_button = Button(root, text="Submit",
                       font=("Alegreya Sans Bold", 11),
                       fg = '#000000', bg = 'white',
                       command = Encode)
submit_ID_button.pack()
submit_ID_button.place(anchor = "w", relx = 0.7, y = 280,
                    height = 25, width = 90)


clear_button = Button(root, text="Clear Text",
                      font=("Alegreya Sans Bold", 11),
                      fg = '#000000', bg = 'white',
                      command = clear_text)
clear_button.pack()
clear_button.place(anchor = "w", relx = 0.60, y = 280,
                   height = 25, width = 90) 

data_label = Label(root, text = "Data Size: " + str(data_length), 
                   font=("Alegreya Sans Bold", 10),
                   bg='#34A853', fg='#FFFFFF')
data_label.pack()
data_label.place(anchor = "w", relx = 0.8, y = 280)

max_data_label = Label(root, text = "Max Data: " + str(max_data),
                       font=("Alegreya Sans Bold", 10),
                       bg='#34A853', fg='#FFFFFF')

max_data_label.pack()
max_data_label.place(anchor = "w", relx = 0.8, y = 300)
    

#vid = cv2.VideoCapture(0)
#vid.release()
#ret, frame = vid.get_frame()

        
#photo = PIL.ImageTk.PhotoImage(image = PIL.Image.fromarray(frame))
#canvas_vid.create_image(0, 0, image = self.photo, anchor = 'nw')


#update()

Display_Data_Size()
root.mainloop()

