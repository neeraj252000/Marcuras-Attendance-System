#C:\ProgramData\Anaconda3\lib\site-packages\cv2\__init__.py

import tkinter as tk
import cv2
import PIL.Image, PIL.ImageTk
import time
import face_recognition
import pickle
import openpyxl
import datetime
import csv
import os

'''
import cv2
print(cv2.__file__)

'''

HIGHT = 1080
WIDTH = 1920

#Loading encoding file
folder_path = os.path.abspath('.')
encoding_file = folder_path + "\\encodings.pickle"
image_read_path = folder_path + "\\stored_photo.jpg"
names_file_path = folder_path + "\\names.csv"
excel_file_path = folder_path + "\\myxl.xlsx"
data = pickle.loads(open(encoding_file, "rb").read())
detected_name = "NULL"

class window:
    def __init__(self,window):
        self.window = window

        window.state('zoomed')        # Displays Full screen application window
        #Canvas
        self.canvas = tk.Canvas(window, height = HIGHT, width = WIDTH, bg = "blue")
        self.canvas.pack()

        #Frame_1 - Heading
        self.frame_1 = tk.Frame(window,bg = '#99ffeb',bd = 5)
        self.frame_1.place(relx=0.5, rely=0.05, relwidth = 0.75, relheight = 0.1, anchor='n')

        #Main label - title
        self.label = tk.Label(self.frame_1, text = "MARCURAS ATTENDANCE", font = ("Times New Roman Bold",40), bg = "#99ffeb", fg = "black", bd = 10)
        self.label.place(relx = 0.5, relwidth = 0.9, relheight= 0.9, anchor='n')

        #Frame_2 - For showing video and output
        self.frame_2 = tk.Frame(window,bg = 'blue')
        self.frame_2.place(relx=0.5, rely=0.2, relwidth = 0.9, relheight = 0.7, anchor = 'n')

        #Frame_3 - Capture button
        self.frame_3 = tk.Frame(window, bg = '#99ffeb')
        self.frame_3.place(relx=0.25, rely=0.8, relwidth = 0.5, relheight = 0.1, anchor = 'n')

        #Fram_6 - Save Button
        self.frame_6 = tk.Frame(window, bg = '#99ffeb')
        self.frame_6.place(relx=0.75, rely = 0.8, relwidth = 0.5, relheight=0.1, anchor = 'n')

        #Frame_4 - For Status Bar
        self.frame_4 = tk.Frame(window, bg = 'White')
        self.frame_4.place(relx =0.5, rely = 0.9, relwidth = 0.6, relheight = 0.05, anchor = 'n')
    
        #Frame_5 - Manual Entry
        self.frame_5 = tk.Frame(window,  bg = 'blue')
        self.frame_5.place(relx = 0.9, rely = 0.2, relwidth = 0.2, relheight = 0.1, anchor = 'n')
    
        self.vid = MyVideoCapture(0)

        #canvas for video 
        self.canvas_vid = tk.Canvas(self.frame_2, width = self.vid.width, height = self.vid.height, bg = 'blue')
        self.canvas_vid.place(relx = 0, rely = 0 ,relwidth =0.5,relheight=1)

        #canvas for result
        self.canvas_res = tk.Canvas(self.frame_2, width = self.vid.width, height = self.vid.height, bg = 'blue')
        self.canvas_res.place(relx = 0.5, rely = 0 ,relwidth =0.5,relheight=1)

        #Button for capturing image
        button = tk.Button(self.frame_3, text= "Capture", font = ("Times New Roman Bold",20),bd = 5, bg = '#d9d9d9', fg="black", command = self.snapshot)
        button.place(relx = 0.5, relwidth=0.5, relheight=1, anchor='n')

        #Button for saving info
        button = tk.Button(self.frame_6, text= "Save", font = ("Times New Roman Bold",20),bd = 5, bg = '#d9d9d9', fg="black", command = self.save)
        button.place(relx = 0.5, relwidth=0.5, relheight=1, anchor='n')

        #Button for Manual Entry
        button = tk.Button(self.frame_5, text= "Manual Entry", font = ("Times New Roman",20),bd = 5, bg = '#d9d9d9', fg="black", command = self.entry)
        button.place(relwidth=1, relheight=1)

        #Copying employee data to attendance file
        reader = csv.reader(open(names_file_path))
        
        #Get month as integer 
        now = datetime.datetime.now()
        month = now.month
        month = int(month)

        x1 = openpyxl.load_workbook(excel_file_path)
        s1 = x1['Sheet'+str(month)]
        max_row = 1

        for row in reader:            
            # print(row[1])
            s1.cell(row = max_row+1,column = 1).value = row[0]
            s1.cell(row = max_row+1,column = 2).value = row[1]
            max_row = max_row+1

        x1.save(excel_file_path)

        self.delay = 10
        self.update()
    

    def update_result(self):
        img_stored = cv2.imread(image_read_path)
        self.photo_res = PIL.ImageTk.PhotoImage(image = PIL.Image.fromarray(img_stored))
        self.canvas_res.create_image(0, 0, image = self.photo_res, anchor = 'nw')

    def detect(self):    
        # global flag_pressed
        data = pickle.loads(open(encoding_file, "rb").read())
        image = cv2.imread(image_read_path)
        #cv2.imshow("Captured_Image", image);
        rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

        # detect the (x, y)-coordinates of the bounding boxes corresponding
        # to each face in the input image, then compute the facial embeddings
        # for each face
        print("[INFO] recognizing faces...")
        boxes = face_recognition.face_locations(rgb, model="hog")
        encodings = face_recognition.face_encodings(rgb, boxes)

        # initialize the list of names for each face detected
        names = []

        # loop over the facial embeddings
        for encoding in encodings:
            # attempt to match each face in the input image to our known
            # encodings
            matches = face_recognition.compare_faces(data["encodings"],
                encoding)
            name = "Unknown"

            # check to see if we have found a match
            if True in matches:
                # find the indexes of all matched faces then initialize a
                # dictionary to count the total number of times each face
                # was matched
                matchedIdxs = [i for (i, b) in enumerate(matches) if b]
                counts = {}

                # loop over the matched indexes and maintain a count for
                # each recognized face face
                for i in matchedIdxs:
                    name = data["names"][i]
                    counts[name] = counts.get(name, 0) + 1

                # determine the recognized face with the largest number of
                # votes (note: in the event of an unlikely tie Python will
                # select first entry in the dictionary)
                name = max(counts, key=counts.get)
            
            # update the list of names
            names.append(name)
            

        # loop over the recognized faces
        for ((top, right, bottom, left), name) in zip(boxes, names):
            # draw the predicted face name on the image
            cv2.rectangle(image, (left, top), (right, bottom), (0, 255, 0), 2)
            y = top - 15 if top - 15 > 15 else top + 15
            cv2.putText(image, name, (left, y), cv2.FONT_HERSHEY_SIMPLEX,
                0.75, (0, 255, 0), 2)

        if names:
            #Prompt for user detected
            label_1 = tk.Label(self.frame_4, text = "Welcome "+names[0] + "!", font = ("Times New Roman Bold",20), bg = "#99ffeb", fg = "black", bd = 10)
            label_1.place(relx = 0.5, relwidth = 0.9, relheight= 0.9, anchor='n')


        elif not names:
            label_2 = tk.Label(self.frame_4, text = "Try Again!" , font = ("Times New Roman Bold",20), bg = "#99ffeb", fg = "red", bd = 10)
            label_2.place(relx = 0.5, relwidth = 0.9, relheight= 0.9, anchor='n')

        
        #Saving the detected image 
        cv2.imwrite(image_read_path, image)    
        
        try:
            return names[0]
        except IndexError:
            return 0

    def snapshot(self):
        # Get a frame from the video source
        ret, frame = self.vid.get_frame()
        if ret:
            cv2.imwrite(image_read_path, frame)
        self.update_result()  
        global detected_name 
        detected_name = self.detect()
        self.update_result()

    def get_detected_ID(self,name):
        #Importing User data CSV file
        reader = csv.reader(open(names_file_path))
        #Creating dictionary with names and IDs
        result = {}
        for row in reader:
            result[row[1]] = row[0]
            #print(row[1])

        ID = result.get(name)
        #Typecasting to int
        #handling error if name not found in dictionary - returning 0
        try:
            return(int(ID))
        except TypeError:
            return 0
        
    def search_entered_ID(self, ID):
        #Importing User data CSV file
        reader = csv.reader(open(names_file_path))
        #Checking for ID in database
        count = 0
        for row in reader:
            count = count + 1
            print(row[0])
            #print(count)
            if(row[0] == str(ID)):
                print(row[1])
                return row[1]

    def close(self):
        self.window.destroy()

    def save(self):
        #Accessing global variable
        global detected_name
        #Get current time as string
        now = datetime.datetime.now()
        current_time = now.strftime("%H:%M:%S")
        #Get date as integer
        date = now.day
        date = int(date)
        # date = 3
        #Get month as integer 
        month = now.month
        month = int(month)

        #Get ID of detected User - detected_name is a global variable
        ID = self.get_detected_ID(detected_name)
        x1 = openpyxl.load_workbook(excel_file_path)
        s1 = x1['Sheet'+str(month)]
        for i in range(1,2000):
            if(s1.cell(row = i,column = 1).value == str(ID)):
                print(s1.cell(row = i,column = 1).value)
                break
        if(i>=1999):
            ID = -1
        if(ID>0):
            cell_check = s1.cell(row = i,column = date*3)
            cell_check = str(cell_check.value)    
            try:
                #Create new window
                confirmation_pop_up = tk.Toplevel(MyWindow)
                if(cell_check == "None" ):
                    s1.cell(row = i,column = date*3).value = current_time
                    #Information displayed post entry
                    canvas = tk.Canvas(confirmation_pop_up, height = 250, width = 500, bg = "#99ffeb")
                    label = tk.Label(canvas, text='Attendance Marked!',font = 20)
                    label.place(relx=0.5, rely=0.2,anchor='n')
                    label = tk.Label(canvas, text='Emplayee ID : ' + str(ID) ,font = 20)
                    label.place(relx=0.5, rely=0.3,anchor='n')
                    label = tk.Label(canvas, text='Emplayee Name : ' + str(detected_name) ,font = 20)
                    label.place(relx=0.5, rely=0.3,anchor='n')
                    label = tk.Label(canvas, text='In Time : ' + current_time ,font = 20)
                    label.place(relx=0.5, rely=0.4,anchor='n')
                    canvas.pack()
                    label_2 = tk.Label(self.frame_4, text = "Thank You!" , font = ("Times New Roman Bold",20), bg = "#99ffeb", fg = "red", bd = 10)
                    label_2.place(relx = 0.5, relwidth = 0.9, relheight= 0.9, anchor='n')
                else:
                    s1.cell(row = i,column = date*3 + 1).value = current_time
                    #Information displayed post entry
                    canvas = tk.Canvas(confirmation_pop_up, height = 250, width = 500, bg = "#99ffeb")
                    label = tk.Label(canvas, text='Attendance Marked!',font = 20)
                    label.place(relx=0.5, rely=0.2,anchor='n')
                    label = tk.Label(canvas, text='Emplayee ID : ' + str(ID) ,font = 20)
                    label.place(relx=0.5, rely=0.3,anchor='n')
                    label = tk.Label(canvas, text='Emplayee Name : ' + str(detected_name) ,font = 20)
                    label.place(relx=0.5, rely=0.4,anchor='n')
                    label = tk.Label(canvas, text='Out Time : ' + current_time ,font = 20)
                    label.place(relx=0.5, rely=0.5,anchor='n')
                    canvas.pack()
                    label_2 = tk.Label(self.frame_4, text = "Thank You!" , font = ("Times New Roman Bold",20), bg = "#99ffeb", fg = "red", bd = 10)
                    label_2.place(relx = 0.5, relwidth = 0.9, relheight= 0.9, anchor='n')
                x1.save(excel_file_path)

                #Exit button
                button = tk.Button(canvas, text = 'Exit', font=20, command = self.close)
                button.place(relx=0.5, rely=0.7, relwidth = 0.4, anchor='n')

                #Closing all windows on Pop Up Close Command
                confirmation_pop_up.protocol("WM_DELETE_WINDOW", self.close)

            #Handling permission error if excell is Open
            except PermissionError:
                label_2 = tk.Label(self.frame_4, text = "Please close the excell sheet before saving!" , font = ("Times New Roman Bold",20), bg = "#99ffeb", fg = "red", bd = 10)
                label_2.place(relx = 0.5, relwidth = 0.9, relheight= 0.9, anchor='n')
        
        else:
            label_2 = tk.Label(self.frame_4, text = "ID not Found!" , font = ("Times New Roman Bold",20), bg = "#99ffeb", fg = "red", bd = 10)
            label_2.place(relx = 0.5, relwidth = 0.9, relheight= 0.9, anchor='n')

    def save_1(self, frame, ID):
        #Get current time as string
        now = datetime.datetime.now()
        current_time = now.strftime("%H:%M:%S")
        #Get date as integer
        date = now.day
        date = int(date)
    
        #Get month as integer 
        month = now.month
        month = int(month)

        #typecasting ID as integer
        ID = int(ID)
        #Get name of entered ID from database
        name = self.search_entered_ID(ID)

        x1 = openpyxl.load_workbook(excel_file_path)
        s1 = x1['Sheet'+ str(month)]
        for i in range(1,2000):
            if(s1.cell(row = i,column = 1).value == str(ID)):
                print(s1.cell(row = i,column = 1).value)
                break
        if(i >= 1999):
            ID = -1
        #Cheking if the entered ID is valid  and is present in the database
        if(name != None and ID>0):
            cell_check = s1.cell(row = i,column = date*3)
            cell_check = str(cell_check.value)    
            try:
                #Create new window
                confirmation_pop_up = tk.Toplevel(MyWindow)
                if(cell_check == "None"):
                    s1.cell(row = i,column = date*3).value = current_time
                #Information displayed post entry
                    canvas = tk.Canvas(confirmation_pop_up, height = 250, width = 500, bg = "#99ffeb")
                    canvas.pack()
                    label = tk.Label(canvas, text='Attendance Marked!',font = 20)
                    label.place(relx=0.5, rely=0.2,anchor='n')
                    label = tk.Label(canvas, text='Emplayee ID : ' + str(ID) ,font = 20)
                    label.place(relx=0.5, rely=0.3,anchor='n')
                    label = tk.Label(canvas, text='Emplayee Name : ' + name ,font = 20)
                    label.place(relx=0.5, rely=0.4,anchor='n')
                    label = tk.Label(canvas, text='In Time : ' + current_time ,font = 20)
                    label.place(relx=0.5, rely=0.5,anchor='n')
                    label_2 = tk.Label(frame, text = "Thank You!" , font = ("Times New Roman Bold",20), bg = "#99ffeb", fg = "red", bd = 10)
                    label_2.place(relx = 0.5, relwidth = 0.9, relheight= 0.9, anchor='n')
                else:
                    s1.cell(row = i,column = date*3 + 1).value = current_time
                #Information displayed post entry
                    canvas = tk.Canvas(confirmation_pop_up, height = 250, width = 500, bg = "#99ffeb")
                    canvas.pack()
                    label = tk.Label(canvas, text='Attendance Marked!',font = 20)
                    label.place(relx=0.5, rely=0.2,anchor='n')
                    label = tk.Label(canvas, text='Emplayee ID : ' + str(ID) ,font = 20)
                    label.place(relx=0.5, rely=0.3,anchor='n')
                    label = tk.Label(canvas, text='Emplayee Name : ' + name ,font = 20)
                    label.place(relx=0.5, rely=0.4,anchor='n')
                    label = tk.Label(canvas, text='Out Time : ' + current_time ,font = 20)
                    label.place(relx=0.5, rely=0.5,anchor='n')
                    label_2 = tk.Label(frame, text = "Thank You!" , font = ("Times New Roman Bold",20), bg = "#99ffeb", fg = "red", bd = 10)
                    label_2.place(relx = 0.5, relwidth = 0.9, relheight= 0.9, anchor='n') 
                x1.save(excel_file_path)
                button = tk.Button(canvas, text = 'Exit', font=20, command = self.close)
                button.place(relx=0.5, rely=0.7, relwidth = 0.4, anchor='n')
                confirmation_pop_up.protocol("WM_DELETE_WINDOW", self.close)

            #Handling permission error if excell is Open
            except PermissionError:
                label = tk.Label(frame,font = 25, text= "Please close the Excell file before saving")
                label.place(relheight=1,relwidth = 1)
        else:
            label = tk.Label(frame,font = 35, text= "Invalid ID!", fg = "red")
            label.place(relheight=1,relwidth = 1)

    def get_input(self,text_object,canvas):
        #Reading text input from text box
        ans = text_object.get()
        #Creating frame for displaying Prompt Message
        frame = tk.Frame(canvas, bg = "#99ffeb")
        frame.place(relx = 0.5,rely = 0.5, relheight = 0.1,relwidth = 0.3, anchor ='n')
        #Checking if the entered number is integer
        #Handling value_error for non-integer entry
        try:
            int(ans)
            label = tk.Label(frame,font = 25, text= "Searching for ID...")
            label.place(relheight=1,relwidth = 1)
            self.save_1(frame,int(ans))
        except ValueError:
            label = tk.Label(frame,font = 25, text= "Please enter a valid number", fg = "red")
            label.place(relheight=1,relwidth = 1)

    def entry(self):
        #Creating new window for Manual_Entry
        newWindow = tk.Toplevel(MyWindow)
        newWindow.state('zoomed')        # Displays Full screen application window
        #Canvas for new_entry window
        canvas = tk.Canvas(newWindow, height = HIGHT, width = WIDTH, bg = "#99ffeb")
        canvas.pack()

        #Title
        label = tk.Label(canvas, text = "Enter Employee ID", font = ("Times New Roman Bold",25), bg = "#99ffeb", fg = "black")
        label.place(rely = 0.1, relx = 0.5, relwidth = 0.4, relheight= 0.1, anchor='n')
        #Creating text entry text box for ID
        textExample=tk.Entry(canvas,font = 25)
        textExample.place(relx = 0.5,rely = 0.3, relheight = 0.05, anchor ='n')
        #Button for submitting ID
        btnRead=tk.Button(canvas, width=10, text="Submit", font = 30 ,command = lambda : self.get_input(textExample,canvas))
        btnRead.place(rely=0.4,relx=0.5,anchor = 'n')

class MyVideoCapture:
    def __init__(self, video_source=0):
        # Open the video source
        self.vid = cv2.VideoCapture(video_source)
        if not self.vid.isOpened():
            raise ValueError("Unable to open video source", video_source)

        # Get video source width and height
        self.width = self.vid.get(cv2.CAP_PROP_FRAME_WIDTH)
        self.height = self.vid.get(cv2.CAP_PROP_FRAME_HEIGHT)

    def get_frame(self):
        if self.vid.isOpened():
            ret, frame = self.vid.read()
            if ret:
                # Return a boolean success flag and the current frame converted to BGR
                return (ret, cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
            else:
                return (ret, None)
        else:
            return (ret, None)

    # Release the video source when the object is destroyed
    def __del__(self):
        if self.vid.isOpened():
            self.vid.release()

#Create a window and pass it to the Application object
MyWindow = tk.Tk()
W1 = window(MyWindow)
MyWindow.mainloop()

